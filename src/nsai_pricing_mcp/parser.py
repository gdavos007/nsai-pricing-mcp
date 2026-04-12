"""
Parser for the NSAI pricing spreadsheet.

NSAI's Excel file typically contains two primary sections / sheets:
  - Monthly Index Prices (spot prices + 12-month rolling averages)
  - First-Day-of-Month Prices (used to compute SEC benchmark prices)

Indices typically tracked:
  Oil:  WTI (West Texas Intermediate), Brent
  Gas:  Henry Hub (HH), Waha, El Paso Natural Gas
  NGL:  Mont Belvieu (propane, butane, ethane, etc.)

Because NSAI occasionally reshuffles columns or renames sheets across updates,
this parser inspects the actual structure rather than assuming fixed positions.
"""

import re
from datetime import date, datetime
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------

def parse_nsai_spreadsheet(path: Path) -> dict[str, dict]:
    """
    Parse all sheets in the NSAI pricing spreadsheet.

    Returns a dict keyed by sheet name.  Each value is:
    {
        "headers": [str, ...],
        "records": [{header: value, ...}, ...],
        "sheet_type": "monthly_index" | "first_day" | "unknown"
    }
    """
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    result: dict[str, dict] = {}

    for name in wb.sheetnames:
        ws: Worksheet = wb[name]
        result[name] = _parse_sheet(ws, name)

    wb.close()
    return result


# ---------------------------------------------------------------------------
# Sheet parsing helpers
# ---------------------------------------------------------------------------

def _parse_sheet(ws: Worksheet, sheet_name: str) -> dict:
    """Parse a single worksheet into a structured dict."""
    raw_rows = list(ws.iter_rows(values_only=True))
    if not raw_rows:
        return {"headers": [], "records": [], "sheet_type": "unknown"}

    # Find the first row that looks like a header (has 2+ text cells)
    header_idx = _find_header_row(raw_rows)
    if header_idx is None:
        return {"headers": [], "records": [], "sheet_type": "unknown"}

    headers = _clean_headers(raw_rows[header_idx])
    records: list[dict] = []

    for row in raw_rows[header_idx + 1:]:
        if all(cell is None for cell in row):
            continue  # Skip blank rows
        record = _build_record(headers, row)
        if _is_meaningful_record(record):
            records.append(record)

    sheet_type = _classify_sheet(sheet_name, headers)

    return {
        "headers": headers,
        "records": records,
        "sheet_type": sheet_type,
    }


def _find_header_row(rows: list[tuple]) -> int | None:
    """Return the index of the first row with ≥2 non-None text cells."""
    for i, row in enumerate(rows[:20]):  # Only scan first 20 rows
        text_cells = [c for c in row if c is not None and isinstance(c, str) and c.strip()]
        if len(text_cells) >= 2:
            return i
    return None


def _clean_headers(row: tuple) -> list[str]:
    """Normalise a header row into clean string labels."""
    headers: list[str] = []
    seen: dict[str, int] = {}

    for cell in row:
        if cell is None:
            label = ""
        else:
            # Collapse whitespace, strip newlines (openpyxl preserves them)
            label = re.sub(r"\s+", " ", str(cell)).strip()

        # De-duplicate column names by appending a counter
        if label in seen:
            seen[label] += 1
            label = f"{label}_{seen[label]}"
        else:
            seen[label] = 0

        headers.append(label)

    return headers


def _build_record(headers: list[str], row: tuple) -> dict[str, Any]:
    """Map header names to cell values, normalising types."""
    record: dict[str, Any] = {}

    for header, cell in zip(headers, row):
        if not header:
            continue  # Skip columns with no header
        record[header] = _normalise_cell(cell)

    return record


def _normalise_cell(cell: Any) -> Any:
    """Convert openpyxl cell values to JSON-safe Python types."""
    if cell is None:
        return None
    if isinstance(cell, (datetime, date)):
        return cell.strftime("%Y-%m-%d")
    if isinstance(cell, float):
        # Round to 4 decimal places to avoid floating-point noise
        return round(cell, 4)
    if isinstance(cell, int):
        return cell
    if isinstance(cell, str):
        return cell.strip() if cell.strip() else None
    return str(cell)


def _is_meaningful_record(record: dict[str, Any]) -> bool:
    """Return True if the record has at least 2 non-None, non-empty values."""
    non_null = sum(1 for v in record.values() if v is not None)
    return non_null >= 2


def _classify_sheet(name: str, headers: list[str]) -> str:
    """Guess the semantic type of a sheet based on its name and headers."""
    name_lower = name.lower()
    header_text = " ".join(h.lower() for h in headers)

    if any(kw in name_lower for kw in ("first", "fdm", "sec", "1st")):
        return "first_day"
    if any(kw in name_lower for kw in ("monthly", "index", "spot", "month")):
        return "monthly_index"

    # Fall back to header heuristics
    if "first" in header_text and "month" in header_text:
        return "first_day"
    if any(kw in header_text for kw in ("rolling", "average", "avg")):
        return "monthly_index"

    return "unknown"


# ---------------------------------------------------------------------------
# Analytical helpers (used by MCP tools)
# ---------------------------------------------------------------------------

def get_latest_record(sheet_data: dict) -> dict | None:
    """Return the most recent record from a parsed sheet."""
    records = sheet_data.get("records", [])
    return records[-1] if records else None


def filter_by_year(sheet_data: dict, year: int) -> list[dict]:
    """Return records whose date column matches the given year."""
    records = sheet_data.get("records", [])
    return [
        r for r in records
        if any(str(v).startswith(str(year)) for v in r.values() if v is not None)
    ]


def filter_by_date_range(
    sheet_data: dict,
    start: str,
    end: str | None = None,
) -> list[dict]:
    """
    Return records within [start, end].
    start / end should be ISO date strings like '2022-01-01' or '2022-01'.
    """
    records = sheet_data.get("records", [])
    result = []

    for record in records:
        # Find the first date-like value in the record
        record_date: str | None = None
        for val in record.values():
            if val and isinstance(val, str) and re.match(r"\d{4}-\d{2}", val):
                record_date = val
                break

        if record_date is None:
            continue
        if record_date < start:
            continue
        if end and record_date > end:
            continue
        result.append(record)

    return result


def summarise_sheet(sheet_data: dict, sheet_name: str) -> dict:
    """Build a concise summary of a sheet for reporting."""
    records = sheet_data.get("records", [])
    return {
        "sheet_name": sheet_name,
        "sheet_type": sheet_data.get("sheet_type", "unknown"),
        "column_count": len(sheet_data.get("headers", [])),
        "row_count": len(records),
        "first_record": records[0] if records else None,
        "latest_record": records[-1] if records else None,
        "headers": sheet_data.get("headers", []),
    }
